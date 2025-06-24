
from flask import Flask, render_template, request, redirect
import cv2
import os
from openpyxl import Workbook, load_workbook
import webbrowser
import threading
import datetime

app = Flask(__name__)

# Ensure image folder exists
IMAGE_FOLDER = "static/images"
os.makedirs(IMAGE_FOLDER, exist_ok=True)

# Excel setup
EXCEL_FILE = "data.xlsx"
if not os.path.exists(EXCEL_FILE):
    wb = Workbook()
    ws = wb.active
    ws.append(["Name", "Email", "Phone", "Image Name", "Timestamp"])
    wb.save(EXCEL_FILE)

@app.route('/')
def index():
    return render_template('form.html')

@app.route('/submit', methods=['POST'])
def submit():
    name = request.form['name']
    email = request.form['email']
    phone = request.form['phone']

    # Start webcam
    cap = cv2.VideoCapture(0)
    if not cap.isOpened():
        return "Webcam could not be opened. Make sure it's connected and not in use."

    ret, frame = cap.read()
    cap.release()

    if not ret:
        return "Failed to capture image."

    timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
    image_name = f"{name}_{timestamp}.png"
    image_path = os.path.join(IMAGE_FOLDER, image_name)
    cv2.imwrite(image_path, frame)

    # Save to Excel
    wb = load_workbook(EXCEL_FILE)
    ws = wb.active
    ws.append([name, email, phone, image_name, timestamp])
    wb.save(EXCEL_FILE)

    return f"""
    <h2>Thank you, {name}!</h2>
    <p>Your image has been saved and data recorded.</p>
    <img src="/{image_path}" width="300"><br>
    <a href="/">Go Back</a>
    """

# Automatically open browser
def open_browser():
    webbrowser.open_new("http://127.0.0.1:5000/")

if __name__ == '_main_':
    threading.Timer(1.5, open_browser).start()
    app.run(debug=True)